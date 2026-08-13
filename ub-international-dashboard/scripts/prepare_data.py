"""Convert UB agreement/activity Excel workbooks into static JSON datasets."""
from __future__ import annotations
import argparse, json, re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
SOURCE=ROOT/'data-source'
OUT=ROOT/'public'/'data'
TODAY=date.today()

CATEGORY_ALIASES={'student exchange':'Student Mobility','mobility':'Student Mobility','research':'Joint Research','guest lecturer':'Guest Lecture','international seminar':'Joint Seminar','conference':'Joint Conference','workshop':'Joint Workshop','community service':'Joint Community Service'}
COUNTRY_ALIASES={'usa':'United States','united states of america':'United States','america! us':'United States','uk':'United Kingdom','russian federation':'Russia','viet nam':'Vietnam','korea, republic of':'South Korea','turkiye':'Türkiye','china ':'China','thailand ':'Thailand'}
VALID_CATEGORIES=['Student Mobility','Joint Research','Visiting Lecturer','Joint Publication','Joint Supervision','Guest Lecture','Double Degree','Joint Degree','Staff Mobility','Joint International Scientific Event','Joint Community Service','Internship','Summer School','Winter School','COIL','Visiting Researcher','Adjunct Professor','Joint Workshop','Joint Seminar','Joint Conference','Research Grant','Staff Training','Curriculum Development','Academic Benchmarking','Laboratory Collaboration','Innovation or Startup Collaboration','Student Competition','Cultural Exchange','Other']

def text(v): return '' if v is None else str(v).strip()
def iso(v):
    if isinstance(v,datetime): return v.date().isoformat()
    if isinstance(v,date): return v.isoformat()
    if isinstance(v,str) and v.strip():
        for fmt in ('%Y-%m-%d','%d/%m/%Y','%d %b %Y','%m/%d/%Y'):
            try:return datetime.strptime(v.strip(),fmt).date().isoformat()
            except ValueError:pass
    return ''
def as_date(v):
    s=iso(v)
    try:return date.fromisoformat(s) if s else None
    except ValueError:return None
def number(v):
    try:return float(v or 0)
    except (TypeError,ValueError):return 0
def integer(v): return int(number(v))
def split_multi(v): return [x.strip() for x in text(v).split('|') if x.strip()]
def country(v):
    s=' '.join(text(v).split()); return COUNTRY_ALIASES.get(s.lower(),s)
def category(v):
    s=' '.join(text(v).split()); return s if s in VALID_CATEGORIES else CATEGORY_ALIASES.get(s.lower(),s or 'Other')
def slug(v):
    s=re.sub(r'[^a-z0-9]+','-',text(v).lower()).strip('-'); return s or 'unknown'
def rows(path,sheet=None):
    wb=load_workbook(path,data_only=True,read_only=True)
    ws=wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    it=ws.iter_rows(values_only=True); headers=[text(x) for x in next(it)]
    for values in it:
        rec=dict(zip(headers,values))
        if any(v not in (None,'') for v in values): yield rec

def duplicate_values(items,key):
    c=Counter(text(x.get(key)) for x in items if text(x.get(key))); return sorted(k for k,v in c.items() if v>1)

def prepare(agreement_file,activity_file):
    quality={'generated_at':datetime.now().isoformat(timespec='seconds'),'agreement':defaultdict(list),'activity':defaultdict(list),'counts':{}}
    mous=[]
    for r in rows(agreement_file):
        mid=text(r.get('MoU ID')); start=as_date(r.get('Start Date')); end=as_date(r.get('End Date'))
        remaining=(end-TODAY).days if end else None
        status='Active' if end and end>=TODAY else 'Expired' if end else 'Unknown'
        item={'mou_id':mid,'mou_name':text(r.get('MoU Name')),'agreement_number':text(r.get('No')),'partner':text(r.get('Partner')),'country':country(r.get('Country')),'scheme':text(r.get('Collaboration Scheme')),'start_date':start.isoformat() if start else '','end_date':end.isoformat() if end else '','duration':text(r.get('Duration')),'status':status,'remaining_days':remaining,'start_year':start.year if start else None,'end_year':end.year if end else None,'document_url':text(r.get('Link Doc')),'expiring_30':remaining is not None and 0<=remaining<=30,'expiring_90':remaining is not None and 0<=remaining<=90,'expiring_180':remaining is not None and 0<=remaining<=180}
        if not start or not end: quality['agreement']['invalid_dates'].append(mid)
        elif end<start: quality['agreement']['invalid_date_ranges'].append(mid)
        if not item['document_url']: quality['agreement']['missing_document_links'].append(mid)
        mous.append(item)

    mou_by_id={m['mou_id']:m for m in mous if m['mou_id']}
    activities=[]
    for r in rows(activity_file,'Activities'):
        aid=text(r.get('Activity ID')); title=text(r.get('Activity Title'))
        if not aid and not title: continue
        start=as_date(r.get('Start Date')); end=as_date(r.get('End Date')) or start
        availability=text(r.get('Agreement Availability')) or ('With Agreement' if text(r.get('MoU ID')) else 'Without Agreement')
        mid=text(r.get('MoU ID')); linked=mou_by_id.get(mid)
        current=text(r.get('Activity Status'))
        if start and end:
            current='Completed' if end<TODAY else 'Planned' if start>TODAY else 'Ongoing'
        item={'activity_id':aid,'title':title,'description':text(r.get('Activity Description')),'category':category(r.get('Activity Category')),'status':current or 'Planned','start_date':start.isoformat() if start else '','end_date':end.isoformat() if end else '','year':start.year if start else None,'duration_days':(end-start).days+1 if start and end else None,'agreement_availability':availability,'reference_type':text(r.get('Agreement Reference Type')),'mou_id':mid,'mou_name':text(r.get('MoU Name')) or (linked or {}).get('mou_name',''),'agreement_number':text(r.get('Agreement Number')),'leading_faculty':text(r.get('Leading UB Faculty')),'faculties':split_multi(r.get('Involved UB Faculties')),'study_programs':split_multi(r.get('Involved UB Study Programs')),'responsible_unit':text(r.get('Responsible UB Unit')),'ub_pic':text(r.get('UB Person in Charge')),'ub_pic_email':text(r.get('UB PIC Email')),'partner':text(r.get('Partner Institution')),'country':country(r.get('Partner Country')),'partner_pic':text(r.get('Partner Contact Person')),'partner_email':text(r.get('Partner Contact Email')),'mode':text(r.get('Activity Mode')),'location':text(r.get('Activity Location')),'mobility_direction':text(r.get('Mobility Direction')),'international_participants':integer(r.get('International Participant Count')),'international_roles':split_multi(r.get('International Participant Role')),'ub_participants':integer(r.get('UB Participant Count')),'ub_roles':split_multi(r.get('UB Participant Role')),'funding_source':text(r.get('Funding Source')),'funding_amount':number(r.get('Funding Amount')),'currency':text(r.get('Currency')),'expected_output':text(r.get('Expected Output')),'actual_output':text(r.get('Actual Output')),'output_type':text(r.get('Output Type')),'output_url':text(r.get('Output Link')),'evidence_url':text(r.get('Evidence Link')),'sdgs':split_multi(r.get('Related SDGs')),'remarks':text(r.get('Remarks')),'last_updated':iso(r.get('Last Updated')),'has_formal_agreement':availability=='With Agreement' and bool(mid),'agreement_match':bool(linked),'agreement_valid_during_activity':bool(linked and start and end and linked['start_date']<=start.isoformat() and linked['end_date']>=end.isoformat())}
        if availability=='With Agreement' and mid and not linked: quality['activity']['unknown_mou_ids'].append(aid)
        if availability=='Without Agreement': quality['activity']['without_agreements'].append(aid)
        if not item['partner']: quality['activity']['missing_partner'].append(aid)
        if not item['country']: quality['activity']['missing_country'].append(aid)
        if not item['leading_faculty'] and not item['faculties']: quality['activity']['missing_faculties'].append(aid)
        if not item['study_programs']: quality['activity']['missing_study_programs'].append(aid)
        if not item['ub_pic']: quality['activity']['missing_pic'].append(aid)
        if not item['international_participants'] and not item['ub_participants']: quality['activity']['missing_participant_counts'].append(aid)
        if not item['evidence_url']: quality['activity']['missing_evidence'].append(aid)
        if not start: quality['activity']['invalid_dates'].append(aid)
        if start and end and end<start: quality['activity']['invalid_date_ranges'].append(aid)
        if item['category'] not in VALID_CATEGORIES: quality['activity']['inconsistent_categories'].append(aid)
        activities.append(item)

    quality['agreement']['duplicate_mou_ids']=duplicate_values(mous,'mou_id')
    quality['agreement']['duplicate_agreement_numbers']=duplicate_values(mous,'agreement_number')
    quality['activity']['duplicate_activity_ids']=duplicate_values(activities,'activity_id')
    by_mou=defaultdict(list)
    for a in activities:
        if a['agreement_match']: by_mou[a['mou_id']].append(a)
    thresholds={'low_max':1,'moderate_max':3}
    for m in mous:
        linked=by_mou[m['mou_id']]; count=len(linked)
        m['activity_count']=count; m['international_participants']=sum(a['international_participants'] for a in linked); m['ub_participants']=sum(a['ub_participants'] for a in linked)
        m['output_count']=sum(bool(a['actual_output'] or a['output_url']) for a in linked); m['faculties']=sorted({f for a in linked for f in ([a['leading_faculty']] if a['leading_faculty'] else [])+a['faculties']}); m['utilization']='No Implementation' if count==0 else 'Low Utilization' if count<=thresholds['low_max'] else 'Moderate Utilization' if count<=thresholds['moderate_max'] else 'High Utilization'
        dates=[a['start_date'] for a in linked if a['start_date']]; m['last_activity_date']=max(dates) if dates else ''

    partner_map={}
    for name in sorted({m['partner'] for m in mous if m['partner']}|{a['partner'] for a in activities if a['partner']}):
        ms=[m for m in mous if m['partner']==name]; acts=[a for a in activities if a['partner']==name]
        partner_map[name]={'partner_id':slug(name),'name':name,'country':next((x['country'] for x in ms+acts if x['country']),''),'agreement_count':len(ms),'active_agreements':sum(m['status']=='Active' for m in ms),'activity_count':len(acts),'with_agreement':sum(a['has_formal_agreement'] for a in acts),'without_agreement':sum(a['agreement_availability']=='Without Agreement' for a in acts),'international_participants':sum(a['international_participants'] for a in acts),'ub_participants':sum(a['ub_participants'] for a in acts),'faculties':sorted({f for a in acts for f in ([a['leading_faculty']] if a['leading_faculty'] else [])+a['faculties']}),'study_programs':sorted({p for a in acts for p in a['study_programs']}),'categories':dict(Counter(a['category'] for a in acts)),'mou_ids':[m['mou_id'] for m in ms],'activity_ids':[a['activity_id'] for a in acts]}
    countries=[]
    for name in sorted({x['country'] for x in mous+activities if x['country']}):
        countries.append({'country':name,'agreements':sum(m['country']==name for m in mous),'activities':sum(a['country']==name for a in activities),'partners':len({m['partner'] for m in mous if m['country']==name}|{a['partner'] for a in activities if a['country']==name})})
    faculties=[]
    fset={f for a in activities for f in ([a['leading_faculty']] if a['leading_faculty'] else [])+a['faculties']}
    for f in sorted(x for x in fset if x): faculties.append({'faculty':f,'activities':sum(f==a['leading_faculty'] or f in a['faculties'] for a in activities),'international_participants':sum(a['international_participants'] for a in activities if f==a['leading_faculty'] or f in a['faculties']),'ub_participants':sum(a['ub_participants'] for a in activities if f==a['leading_faculty'] or f in a['faculties'])})
    summary={'generated_at':datetime.now().isoformat(timespec='seconds'),'as_of_date':TODAY.isoformat(),'total_agreements':len(mous),'active_agreements':sum(m['status']=='Active' for m in mous),'expired_agreements':sum(m['status']=='Expired' for m in mous),'expiring_90':sum(m['expiring_90'] for m in mous),'total_activities':len(activities),'activities_with_agreement':sum(a['has_formal_agreement'] for a in activities),'activities_without_agreement':sum(a['agreement_availability']=='Without Agreement' for a in activities),'ongoing_activities':sum(a['status']=='Ongoing' for a in activities),'completed_activities':sum(a['status']=='Completed' for a in activities),'partner_institutions':len(partner_map),'partner_countries':len(countries),'international_participants':sum(a['international_participants'] for a in activities),'ub_participants':sum(a['ub_participants'] for a in activities),'agreement_utilization_rate':round(100*sum(m['activity_count']>0 for m in mous)/len(mous),1) if mous else 0,'activities_without_agreement_pct':round(100*sum(a['agreement_availability']=='Without Agreement' for a in activities)/len(activities),1) if activities else 0,'average_activities_per_active_mou':round(sum(len(v) for v in by_mou.values())/max(1,sum(bool(v) for v in by_mou.values())),2),'utilization_thresholds':thresholds}
    quality['counts']={f'agreement.{k}':len(v) for k,v in quality['agreement'].items()}|{f'activity.{k}':len(v) for k,v in quality['activity'].items()}
    quality['agreement']=dict(quality['agreement']); quality['activity']=dict(quality['activity'])
    return mous,activities,list(partner_map.values()),countries,faculties,summary,quality

def write_json(name,data):
    OUT.mkdir(parents=True,exist_ok=True); (OUT/name).write_text(json.dumps(data,ensure_ascii=False,separators=(',',':')),encoding='utf-8')

def main():
    p=argparse.ArgumentParser(); p.add_argument('--agreements',type=Path,default=SOURCE/'Database_Kerjasama_Luar_Negeri_master.xlsx'); p.add_argument('--activities',type=Path,default=SOURCE/'Database_Aktivitas_Kerjasama_Internasional.xlsx'); args=p.parse_args()
    data=prepare(args.agreements,args.activities)
    for name,value in zip(['mous.json','activities.json','partners.json','countries.json','faculties.json','summary.json','data-quality.json'],data): write_json(name,value)
    print(json.dumps({'agreements':len(data[0]),'activities':len(data[1]),'partners':len(data[2]),'countries':len(data[3]),'output':str(OUT)},indent=2))
if __name__=='__main__': main()
